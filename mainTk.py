from course_scrap_parse import *
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill

from tkinter import Tk, Frame, Label, Entry, StringVar
from tkinter import LEFT, BOTH, YES, NW, SE
from tkinter.ttk import Combobox, Button
from tkinter.filedialog import asksaveasfilename
from tkinter.messagebox import showinfo, showerror

from datetime import datetime

base_url = "https://w5.ab.ust.hk/wcq/cgi-bin"

semesters = {
    "Fall": "10",
    "Winter": "20",
    "Spring": "30",
    "Summer": "40",
}

class MainWindow(Frame):
    def __init__(self, parent, *args, **kwargs):
        Frame.__init__(self, parent, *args, **kwargs)
        self.parent = parent

        Label(self, text="Year:").grid(row=0, column=0, sticky=NW)
        self.year_combo = Combobox(self, width=6, state="readonly", values=self._get_year())
        self.year_combo.bind("<<ComboboxSelected>>", self._update_url)
        self.year_combo.current(1)
        self.year_combo.grid(row=0, column=1, sticky=NW)

        Label(self, text="Semester:").grid(row=1, column=0, sticky=NW)
        self.sem_combo = Combobox(self, width=8, state="readonly", values=["Fall", "Winter", "Spring", "Summer"])
        self.sem_combo.bind("<<ComboboxSelected>>", self._update_url)
        self.sem_combo.current(0)
        self.sem_combo.grid(row=1, column=1, sticky=NW)

        self.urlVar = StringVar()
        Label(self, text="Website:").grid(row=2, column=0, sticky=NW)
        Entry(self, textvariable=self.urlVar, width=50).grid(row=2, column=1, sticky=NW)

        Button(self, text="Generate Excel", command=self._generate).grid(row=3, column=1, sticky=SE)

        self._update_url(None)

    def _get_year(self):
        cur_year = datetime.now().year
        return [cur_year-1, cur_year, cur_year+1]

    def _update_url(self, selected):
        year = self.year_combo.get()[-2:]
        semester = semesters[self.sem_combo.get()]
        self.urlVar.set(f"{base_url}/{year}{semester}/common_core/4Y")

    def _get_2022_cc_url(self):
        year = self.year_combo.get()[-2:]
        semester = semesters[self.sem_combo.get()]
        return f"{base_url}/{year}{semester}/common_core/CC22"

    def _generate(self):
        filename = asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel File", ".xlsx")], title="Choose Location")
        if not filename:
            return

        # get course HTML components
        url = self.urlVar.get()
        try:
            page = getHtml(url)
            courses = page.findAll("div", {"course"})

            """20/09/2022 Update:

            Since 01/09/2022, the new admitted students can take CORE1401, 1402 and 1404
            to fulfil E-Comm requirement. However, other students cannot take the courses
            to fulfil the requirement. As such the common core website doesn't include the
            courses mentioned above.

            To solve this issue, the workaround solution will be scrapping two different
            common core website:
            1. CC list for students admitted before 2022
            1. CC list for students admitted from 2022
            """

            """course_code = c.findAll('h2')[-1]

            Here we create a set that contains all course code.
            """
            course_book = set(
                [c.findAll('h2')[-1] for c in courses]
            )

            cc22page = getHtml(self._get_2022_cc_url())
            for c in cc22page.findAll("div", {"course"}):
                title = c.findAll('h2')[-1]
                if title not in course_book: # if any course exits in CC22 not in YR4 list, add the coruse 
                    courses.append(c)
                    course_book.add(title)

            # fixing the course order
            courses.sort(key=lambda c: c.findAll('h2')[-1].text)
            # for c in courses:
            #     title = c.findAll('h2')[-1]
            #     print(title.text)

        except Exception as e:
            showerror("Error", f"Unable to retrieve webpage from {url}\n\nError:\n{e}")
            return

        # create workbook with column title
        try:
            wb = Workbook()
            sh = wb[wb.sheetnames[0]]
            sh.title = "Common Core"
            sh.append([
                "Common Core Area",
                "School",
                "Subject Area",
                "Course Number",
                "Course Title",
                "Units",
                "Instructor",
                "Section",
                "Quota for Each Section",
                "Total Quota",
            ])
        except:
            showerror("Error", f"Unable to create Excel file")
            return

        # parse the HTML course components and put into a row in sheet
        error = []
        try:
            for i, course in enumerate(courses, start=2):
                c, success1, err1 = parseCourse(course)
                c, success2, err2 = formatCourse(c)

                for title, content in c.items():
                    if type(content) == list:
                        c[title] = str(content)

                sh.append([
                    c["Common Core Area"],
                    c["School"],
                    c["Subject Area"],
                    c["Course Number"],
                    c["Course Title"],
                    c["Units"],
                    c["Instructor"],
                    c["Section"],
                    c["Quota for Each Section"],
                    c["Total Quota"],
                ])

                if not (success1 and success2):
                    if err1 is not None: error.append(str(err1))
                    if err2 is not None: error.append(str(err2))
                    redFill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
                    for row in sh.iter_rows(min_row=i, max_row=i, min_col=1, max_col=10):
                        for cell in row:
                            cell.fill = redFill
        except Exception as e:
            showerror("Error", f"Unable to parse webpage\n\nError:\n{e}")
            return

        try:
            # excel file formatting
            for col in ["A", "B", "C", "D", "E"]:
                for cell in sh[col]:
                    cell.alignment = Alignment(vertical="center")

            for cell in sh["G"]:
                    cell.alignment = Alignment(vertical="center", wrapText=True)

            for col in ["H", "I"]:
                for i, cell in enumerate(sh[col]):
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrapText=True if i != 0 else False)

            for col in ["F", "J"]:
                for i, cell in enumerate(sh[col]):
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            for column_cells in sh.columns:
                length = max(len(str(cell.value)) if "\n" not in str(cell.value) else max(len(s) for s in str(cell.value).split("\n")) for cell in column_cells[1:])
                sh.column_dimensions[column_cells[0].column_letter].width = length

            # save and done
            wb.save(filename)
        except Exception as e:
            showerror("Error", f"Unable to save Excel file\n\nError:\n{e}")

        showinfo(title="Done", message=f"The Excel file is generated and save at {filename}")
        if len(error):
            err_msg = f"{len(error)} error encountered during the process:" + "\n" + '\n'.join(error)
            showinfo(title="Error", message=err_msg)

if __name__ == '__main__':
    root = Tk()
    root.title("HKUST Course Scrap")
    root.geometry("570x120")

    MainWindow(root).pack(side=LEFT, fill=BOTH, expand=YES)
    root.mainloop()